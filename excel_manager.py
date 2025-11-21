#!/usr/bin/env python3
"""
Excelベースのデータ管理モジュール
スタッフマスタと希望データをExcelファイルで管理
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from pathlib import Path
import calendar

class ExcelManager:
    def __init__(self, base_path='output'):
        self.base_path = Path(base_path)
        self.base_path.mkdir(exist_ok=True)
        self.shift_types = ['早番', '中番', '遅番', '休み', '有給', '半休']

    def create_month_shift(self, year_month, staff_names):
        """
        月次シフトフォルダとスタッフマスタを作成

        Args:
            year_month: 例: "2024年11月"
            staff_names: スタッフ名のリスト
        """
        # フォルダ作成
        month_folder = self.base_path / year_month
        month_folder.mkdir(exist_ok=True)

        # スタッフマスタファイルを作成
        master_file = month_folder / 'スタッフマスタ.xlsx'
        self._create_staff_master(master_file, staff_names)

        return month_folder

    def _create_staff_master(self, file_path, staff_names):
        """スタッフマスタExcelファイルを作成"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'スタッフマスタ'

        # スタイル設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_align = Alignment(horizontal='center', vertical='center')

        # ヘッダー
        headers = ['スタッフ名', '提出状態', '提出日時']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # スタッフデータ
        for row, name in enumerate(staff_names, start=2):
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value='未提出').border = thin_border
            ws.cell(row=row, column=3, value='').border = thin_border
            ws.cell(row=row, column=2).alignment = center_align

        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25

        wb.save(file_path)

    def get_staff_list(self, year_month):
        """スタッフマスタからスタッフ一覧を取得"""
        master_file = self.base_path / year_month / 'スタッフマスタ.xlsx'

        if not master_file.exists():
            return []

        wb = openpyxl.load_workbook(master_file)
        ws = wb.active

        staff_list = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            status = ws.cell(row=row, column=2).value
            submitted_at = ws.cell(row=row, column=3).value

            if name:
                staff_list.append({
                    'name': name,
                    'submitted': status == '提出済み',
                    'submitted_at': submitted_at
                })

        wb.close()
        return staff_list

    def save_staff_preferences(self, year_month, staff_name, preferences_dict):
        """
        スタッフの希望をExcelファイルに保存

        Args:
            year_month: 例: "2024年11月"
            staff_name: スタッフ名
            preferences_dict: {日付: 希望シフト}の辞書
        """
        month_folder = self.base_path / year_month
        pref_file = month_folder / f'希望_{staff_name}.xlsx'

        # 希望ファイルを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '希望シフト'

        # スタイル設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        # ヘッダー
        ws.cell(row=1, column=1, value='日付').font = header_font
        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=2, value='曜日').font = header_font
        ws.cell(row=1, column=2).border = thin_border
        ws.cell(row=1, column=2).alignment = center_align
        ws.cell(row=1, column=3, value='希望シフト').font = header_font
        ws.cell(row=1, column=3).border = thin_border
        ws.cell(row=1, column=3).alignment = center_align

        # データ
        for row, (day, preference) in enumerate(sorted(preferences_dict.items()), start=2):
            # 日付情報を取得
            year, month = self._parse_year_month(year_month)
            date = datetime(year, month, day)
            weekday = ['月', '火', '水', '木', '金', '土', '日'][date.weekday()]

            ws.cell(row=row, column=1, value=f'{month}/{day}').border = thin_border
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2, value=weekday).border = thin_border
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3, value=preference).border = thin_border
            ws.cell(row=row, column=3).alignment = center_align

        # 列幅調整
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15

        wb.save(pref_file)
        wb.close()

        # スタッフマスタを更新
        self._update_staff_master(year_month, staff_name, submitted=True)

    def load_staff_preferences(self, year_month, staff_name):
        """スタッフの希望を読み込み"""
        month_folder = self.base_path / year_month
        pref_file = month_folder / f'希望_{staff_name}.xlsx'

        if not pref_file.exists():
            return {}

        wb = openpyxl.load_workbook(pref_file)
        ws = wb.active

        preferences = {}
        for row in range(2, ws.max_row + 1):
            date_str = ws.cell(row=row, column=1).value
            preference = ws.cell(row=row, column=3).value

            if date_str and preference:
                # "11/15" -> 15
                day = int(date_str.split('/')[-1])
                preferences[day] = preference

        wb.close()
        return preferences

    def _update_staff_master(self, year_month, staff_name, submitted):
        """スタッフマスタの提出状態を更新"""
        master_file = self.base_path / year_month / 'スタッフマスタ.xlsx'

        if not master_file.exists():
            return

        wb = openpyxl.load_workbook(master_file)
        ws = wb.active

        # スタッフを検索して更新
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name == staff_name:
                if submitted:
                    ws.cell(row=row, column=2, value='提出済み')
                    ws.cell(row=row, column=3, value=datetime.now().strftime('%Y/%m/%d %H:%M'))
                else:
                    ws.cell(row=row, column=2, value='未提出')
                    ws.cell(row=row, column=3, value='')
                break

        wb.save(master_file)
        wb.close()

    def check_all_submitted(self, year_month):
        """全員が提出済みかチェック"""
        staff_list = self.get_staff_list(year_month)

        if not staff_list:
            return False

        return all(staff['submitted'] for staff in staff_list)

    def get_all_preferences(self, year_month):
        """全スタッフの希望を取得"""
        staff_list = self.get_staff_list(year_month)
        all_preferences = {}

        for staff in staff_list:
            if staff['submitted']:
                preferences = self.load_staff_preferences(year_month, staff['name'])
                all_preferences[staff['name']] = preferences

        return all_preferences

    def _parse_year_month(self, year_month):
        """年月文字列から年と月を抽出 (例: "2024年11月" -> (2024, 11))"""
        parts = year_month.replace('年', ' ').replace('月', '').split()
        year = int(parts[0])
        month = int(parts[1])
        return year, month

    def get_month_dates(self, year_month):
        """指定された年月の日付情報を取得"""
        year, month = self._parse_year_month(year_month)
        num_days = calendar.monthrange(year, month)[1]

        dates = []
        for day in range(1, num_days + 1):
            date = datetime(year, month, day)
            dates.append({
                'day': day,
                'date': date,
                'weekday': ['月', '火', '水', '木', '金', '土', '日'][date.weekday()],
                'formatted': f"{month}/{day}"
            })

        return dates

    def shift_exists(self, year_month):
        """指定された年月のシフトが存在するかチェック"""
        month_folder = self.base_path / year_month
        master_file = month_folder / 'スタッフマスタ.xlsx'
        return master_file.exists()

    def get_available_months(self):
        """利用可能な月の一覧を取得"""
        months = []
        for folder in self.base_path.iterdir():
            if folder.is_dir():
                master_file = folder / 'スタッフマスタ.xlsx'
                if master_file.exists():
                    months.append(folder.name)
        return sorted(months, reverse=True)

    def add_staff(self, year_month, staff_name):
        """スタッフを追加"""
        master_file = self.base_path / year_month / 'スタッフマスタ.xlsx'

        if not master_file.exists():
            return False

        wb = openpyxl.load_workbook(master_file)
        ws = wb.active

        # 既に存在するかチェック
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name == staff_name:
                wb.close()
                return False  # 既に存在

        # 新しい行を追加
        new_row = ws.max_row + 1
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        ws.cell(row=new_row, column=1, value=staff_name).border = thin_border
        ws.cell(row=new_row, column=2, value='未提出').border = thin_border
        ws.cell(row=new_row, column=2).alignment = center_align
        ws.cell(row=new_row, column=3, value='').border = thin_border

        wb.save(master_file)
        wb.close()
        return True

    def remove_staff(self, year_month, staff_name):
        """スタッフを削除"""
        master_file = self.base_path / year_month / 'スタッフマスタ.xlsx'

        if not master_file.exists():
            return False

        wb = openpyxl.load_workbook(master_file)
        ws = wb.active

        # スタッフを検索して削除
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name == staff_name:
                ws.delete_rows(row)
                wb.save(master_file)
                wb.close()

                # 希望ファイルも削除
                pref_file = self.base_path / year_month / f'希望_{staff_name}.xlsx'
                if pref_file.exists():
                    pref_file.unlink()

                return True

        wb.close()
        return False
