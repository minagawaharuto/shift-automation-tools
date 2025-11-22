#!/usr/bin/env python3
"""
シンプルシフト作成アプリ
スタッフが個別に希望を入力し、全員完了で自動最適化
"""

from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from datetime import datetime
from pathlib import Path
import subprocess
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from excel_manager import ExcelManager

app = Flask(__name__)
app.secret_key = 'simple_shift_app_2024'

# 基本設定
BASE_DIR = Path(__file__).parent
excel_mgr = ExcelManager(base_path='output')
SHIFT_TYPES = ['早番', '中番', '遅番', '休み', '有給', '半休']

# 現在の月を取得
def get_current_month():
    now = datetime.now()
    return f"{now.year}年{now.month}月"


@app.route('/')
def index():
    """トップページ"""
    year_month = get_current_month()
    shift_exists = excel_mgr.shift_exists(year_month)

    # 提出状況を取得
    staff_list = []
    all_submitted = False
    optimized = False

    if shift_exists:
        staff_list = excel_mgr.get_staff_list(year_month)
        all_submitted = excel_mgr.check_all_submitted(year_month)
        output_file = Path('output') / year_month / f"{year_month}_最適化シフト_完成版.xlsx"
        optimized = output_file.exists()

    return render_template('index.html',
                         year_month=year_month,
                         shift_exists=shift_exists,
                         staff_list=staff_list,
                         all_submitted=all_submitted,
                         optimized=optimized)


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """初回設定：スタッフ登録"""
    year_month = get_current_month()

    if request.method == 'POST':
        try:
            # スタッフ名を取得
            staff_names = []
            for key in request.form.keys():
                if key.startswith('staff_name_'):
                    name = request.form[key].strip()
                    if name:
                        staff_names.append(name)

            if len(staff_names) == 0:
                flash('スタッフ名を最低1名入力してください', 'error')
                return redirect(url_for('setup'))

            # 月次シフトを作成
            excel_mgr.create_month_shift(year_month, staff_names)

            flash(f'{year_month}のスタッフ登録が完了しました！', 'success')
            return redirect(url_for('index'))

        except Exception as e:
            flash(f'エラーが発生しました: {str(e)}', 'error')
            return redirect(url_for('setup'))

    # GET: フォーム表示
    return render_template('setup.html', year_month=year_month)


@app.route('/manage_staff', methods=['GET', 'POST'])
def manage_staff():
    """スタッフ管理：追加・削除"""
    year_month = get_current_month()

    if not excel_mgr.shift_exists(year_month):
        flash('まだスタッフ登録されていません。初回設定を行ってください。', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        staff_name = request.form.get('staff_name', '').strip()

        if action == 'add':
            if not staff_name:
                flash('スタッフ名を入力してください', 'error')
            else:
                if excel_mgr.add_staff(year_month, staff_name):
                    flash(f'{staff_name}を追加しました！', 'success')
                else:
                    flash(f'{staff_name}は既に登録されています', 'error')

        elif action == 'remove':
            if not staff_name:
                flash('スタッフ名が不正です', 'error')
            else:
                # 最低1名は残す必要がある
                staff_list = excel_mgr.get_staff_list(year_month)
                if len(staff_list) <= 1:
                    flash('最低1名のスタッフが必要です', 'error')
                else:
                    if excel_mgr.remove_staff(year_month, staff_name):
                        flash(f'{staff_name}を削除しました', 'success')
                    else:
                        flash(f'{staff_name}が見つかりませんでした', 'error')

        return redirect(url_for('manage_staff'))

    # GET: スタッフ一覧を表示
    staff_list = excel_mgr.get_staff_list(year_month)
    return render_template('staff_manage.html', year_month=year_month, staff_list=staff_list)


@app.route('/input/<staff_name>')
def input_form(staff_name):
    """希望入力フォーム"""
    year_month = get_current_month()

    if not excel_mgr.shift_exists(year_month):
        flash('まだスタッフ登録されていません。初回設定を行ってください。', 'error')
        return redirect(url_for('index'))

    # 日付情報を取得
    dates = excel_mgr.get_month_dates(year_month)

    # 既存の希望があれば読み込み
    existing_preferences = excel_mgr.load_staff_preferences(year_month, staff_name)

    return render_template('input.html',
                         year_month=year_month,
                         staff_name=staff_name,
                         dates=dates,
                         shift_types=SHIFT_TYPES,
                         existing_preferences=existing_preferences)


@app.route('/submit', methods=['POST'])
def submit():
    """希望を提出"""
    try:
        year_month = get_current_month()
        staff_name = request.form.get('staff_name')

        if not staff_name:
            flash('スタッフ名が不正です', 'error')
            return redirect(url_for('index'))

        # 希望データを取得
        preferences = {}
        for key, value in request.form.items():
            if key.startswith('shift_'):
                day = int(key.split('_')[1])
                preferences[day] = value

        # データを保存
        excel_mgr.save_staff_preferences(year_month, staff_name, preferences)

        flash(f'{staff_name}さんの希望を提出しました！', 'success')

        # 全員提出済みかチェック
        if excel_mgr.check_all_submitted(year_month):
            flash('全員の提出が完了しました！シフトを最適化しています...', 'info')

            # 自動最適化を実行
            result = run_optimizer(year_month)

            if result['success']:
                flash('シフトの最適化が完了しました！', 'success')
                return redirect(url_for('complete'))
            else:
                flash(f'最適化に失敗しました: {result["error"]}', 'error')

        return redirect(url_for('index'))

    except Exception as e:
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/complete')
def complete():
    """完了画面"""
    year_month = get_current_month()
    output_file = Path('output') / year_month / f"{year_month}_最適化シフト_完成版.xlsx"

    if not output_file.exists():
        flash('最適化されたシフトファイルが見つかりません', 'error')
        return redirect(url_for('index'))

    return render_template('complete.html', year_month=year_month)


@app.route('/download')
def download():
    """最適化されたシフトをダウンロード"""
    year_month = get_current_month()
    output_file = Path('output') / year_month / f"{year_month}_最適化シフト_完成版.xlsx"

    if not output_file.exists():
        flash('ファイルが見つかりません', 'error')
        return redirect(url_for('index'))

    return send_file(
        output_file,
        as_attachment=True,
        download_name=f"{year_month}_最適化シフト.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def run_optimizer(year_month):
    """最適化を実行"""
    try:
        # 全員の希望を統合してExcelファイルを作成
        all_preferences = excel_mgr.get_all_preferences(year_month)
        staff_list = excel_mgr.get_staff_list(year_month)
        staff_names = [s['name'] for s in staff_list]

        if not staff_names:
            return {'success': False, 'error': 'スタッフが登録されていません'}

        # 入力ファイルを作成
        month_folder = Path('output') / year_month
        input_file = month_folder / f"{year_month}.xlsx"
        create_input_excel(input_file, staff_names, year_month, all_preferences)

        # 最適化スクリプトの存在確認
        optimizer_path = BASE_DIR / 'Shift_optimizer.py'
        if not optimizer_path.exists():
            return {'success': False, 'error': f'最適化スクリプトが見つかりません: {optimizer_path}'}

        # Python実行コマンド（現在の環境のPythonを使用）
        python_exe = sys.executable
        
        # 最適化スクリプトを実行
        result = subprocess.run(
            [python_exe, str(optimizer_path), year_month],
            capture_output=True,
            text=True,
            encoding='utf-8',
            cwd=str(BASE_DIR),
            timeout=300  # 5分のタイムアウト
        )

        print("=== Optimizer Output ===")
        stdout_text = result.stdout if result.stdout else ""
        stderr_text = result.stderr if result.stderr else ""
        
        if stdout_text:
            print(stdout_text)
        else:
            print("(標準出力なし)")
        if stderr_text:
            print("=== Optimizer Errors ===")
            print(stderr_text)

        if result.returncode == 0:
            # 出力ファイルの存在確認
            output_file = month_folder / f"{year_month}_最適化シフト_完成版.xlsx"
            if output_file.exists():
                return {'success': True}
            else:
                return {'success': False, 'error': '最適化は完了しましたが、出力ファイルが見つかりません'}
        else:
            # エラーメッセージを詳細化（stdoutとstderrの両方を確認）
            error_msg = ''
            
            # stderrを優先、なければstdoutを使用
            if stderr_text:
                error_msg = stderr_text.strip()
            elif stdout_text:
                error_msg = stdout_text.strip()
            
            # エラーメッセージが見つからない場合
            if not error_msg:
                error_msg = f'最適化スクリプトがエラーコード {result.returncode} で終了しました'
                # stdoutに何か出力されていれば、それを追加情報として含める
                if stdout_text:
                    # 最後の数行を取得（エラーメッセージが含まれている可能性がある）
                    lines = stdout_text.strip().split('\n')
                    if lines:
                        last_lines = '\n'.join(lines[-5:])  # 最後の5行
                        error_msg += f'\n\n出力内容（最後の数行）:\n{last_lines}'
            
            # よくあるエラーパターンをチェック
            if 'ModuleNotFoundError' in error_msg or 'ImportError' in error_msg:
                error_msg += '\n\n（必要なライブラリがインストールされていない可能性があります。ortools、pandas、openpyxlを確認してください）'
            elif 'FileNotFoundError' in error_msg or '見つかりません' in error_msg:
                error_msg += f'\n\n（入力ファイル: {input_file} が正しく作成されているか確認してください）'
            elif 'エラー' in error_msg or 'エラー:' in error_msg:
                # エラーメッセージが既に含まれている場合はそのまま使用
                pass
            elif '最適解が見つかりません' in error_msg or '失敗しました' in error_msg:
                error_msg += '\n\n（制約条件を満たす解が見つかりませんでした。スタッフの希望を調整してください）'
            
            return {'success': False, 'error': error_msg}

    except subprocess.TimeoutExpired:
        return {'success': False, 'error': '最適化処理がタイムアウトしました（5分以上かかっています）'}
    except FileNotFoundError as e:
        return {'success': False, 'error': f'ファイルが見つかりません: {str(e)}'}
    except Exception as e:
        return {'success': False, 'error': f'予期しないエラーが発生しました: {str(e)}'}


def create_input_excel(file_path, staff_names, year_month, all_preferences):
    """統合Excelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "希望シフト"

    # スタイル設定
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # 日付情報を取得
    dates = excel_mgr.get_month_dates(year_month)

    # ヘッダー行
    ws.cell(row=1, column=1, value='氏名').font = header_font
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=1, column=1).alignment = center_align

    for idx, date_info in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=idx, value=date_info['date'])
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
        cell.number_format = 'M/D'

    # スタッフデータ
    for staff_idx, staff_name in enumerate(staff_names, start=2):
        cell = ws.cell(row=staff_idx, column=1, value=staff_name)
        cell.border = thin_border
        cell.alignment = center_align

        preferences = all_preferences.get(staff_name, {})
        for date_idx, date_info in enumerate(dates, start=2):
            day = date_info['day']
            shift_value = preferences.get(day, 'どちらでも')

            cell = ws.cell(row=staff_idx, column=date_idx, value=shift_value)
            cell.border = thin_border
            cell.alignment = center_align

    # 列幅調整
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(dates) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10

    wb.save(file_path)
    print(f"✅ 入力ファイルを作成: {file_path}")


if __name__ == '__main__':
    print("=" * 60)
    print("🚀 シンプルシフト作成アプリ")
    print("=" * 60)
    print(f"📁 作業ディレクトリ: {BASE_DIR}")
    print(f"📁 出力ディレクトリ: output/")
    print("🌐 ブラウザで http://localhost:5000 にアクセス")

    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
