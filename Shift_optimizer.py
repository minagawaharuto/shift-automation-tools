#!/usr/bin/env python3
"""
シフト自動最適化プログラム Ver3.2 (PAD連携対応版)
・PADからの引数受け取り機能を追加
・保存処理の確実性を向上（一時ファイル経由）
・プルダウンメニューを別シート参照方式に変更
"""

import pandas as pd
import numpy as np
from ortools.sat.python import cp_model
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys
import os
import time
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# GUIの利用可能性をチェック
try:
    import tkinter as tk
    from tkinter import simpledialog, messagebox
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False

class ShiftOptimizerV3_2:
    def __init__(self, base_path="."):
        self.base_path = Path(base_path)
        self.shift_types = ['早番', '中番', '遅番', '休み', '有給', '半休']
        self.target_rest_days = 10
        
    def get_year_month_input(self):
        """年月の入力（PAD連携対応）"""
        
        # ★★★ 修正箇所ここから ★★★
        # コマンドライン引数（PADからの入力）があるかチェック
        if len(sys.argv) > 1:
            # 引数を受け取り、余計な空白やクォーテーションを除去して返す
            input_val = sys.argv[1].strip().strip('"').strip("'")
            print(f"🤖 PADから指定された年月: {input_val}")
            return input_val
        # ★★★ 修正箇所ここまで ★★★

        print("\n" + "=" * 60)
        print("🗓️ シフト作成月の指定")
        print("=" * 60)
        
        if GUI_AVAILABLE:
            try:
                root = tk.Tk()
                root.withdraw()
                year_month = simpledialog.askstring(
                    "シフト作成月の指定",
                    "作成したいシフトの年月を入力してください\n（例：2025年11月）",
                    initialvalue=f"{datetime.now().year}年{datetime.now().month}月"
                )
                root.destroy()
                if not year_month: sys.exit(0)
            except:
                year_month = input("作成したいシフトの年月を入力してください（例：2025年11月）: ").strip()
        else:
            year_month = input("作成したいシフトの年月を入力してください（例：2025年11月）: ").strip()
            
        if not year_month: sys.exit(0)
        return year_month
    
    def setup_file_paths(self, year_month):
        folder_path = self.base_path / year_month
        if not folder_path.exists():
            print(f"\n❌ エラー: フォルダが見つかりません: {folder_path}")
            # エラーが見えるように少し待機してから終了
            time.sleep(3)
            sys.exit(1)
        
        print(f"📁 作業フォルダ: {folder_path}")
        input_file = folder_path / f"{year_month}.xlsx"
        # 最終的な出力ファイル名
        output_file = folder_path / f"{year_month}_最適化シフト_完成版.xlsx"
        
        return folder_path, input_file, output_file
    
    def load_excel_from_folder(self, input_file, year_month):
        if not input_file.exists():
            print(f"\n❌ エラー: 入力ファイルが見つかりません: {input_file}")
            time.sleep(3)
            sys.exit(1)
        
        print(f"📂 ファイルを読み込み中: {input_file}")
        # data_only=Trueで値のみ読み込む（数式エラー回避）
        wb = openpyxl.load_workbook(input_file, data_only=True)
        sheet = wb.active
        
        employees = []
        preferences = {}
        date_columns = []
        
        # 日付取得
        for col in range(2, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                if isinstance(cell_value, (int, float)):
                    excel_date = datetime(1899, 12, 30) + timedelta(days=int(cell_value))
                    date_columns.append(excel_date)
                elif isinstance(cell_value, datetime):
                     date_columns.append(cell_value)
        
        # 従業員データ取得
        for row in range(2, sheet.max_row + 1):
            emp_name = sheet.cell(row=row, column=1).value
            if emp_name:
                employees.append(emp_name)
                preferences[emp_name] = []
                for col in range(2, 2 + len(date_columns)):
                    cell_value = sheet.cell(row=row, column=col).value
                    pref_value = str(cell_value) if cell_value else 'どちらでも'
                    # 「中番」は現在の制約では扱えないため「どちらでも」として扱う
                    if pref_value == '中番':
                        pref_value = 'どちらでも'
                    preferences[emp_name].append(pref_value)
        
        wb.close()
        print(f"✅ 読み込み成功: {len(employees)}名, {len(date_columns)}日間")
        return employees, preferences, date_columns
    
    def optimize_shifts(self, employees, preferences, num_days):
        print("\n🔧 最適化を開始...")
        num_employees = len(employees)
        num_shifts = 3
        
        model = cp_model.CpModel()
        shifts = {}
        for e in range(num_employees):
            for d in range(num_days):
                for s in range(num_shifts):
                    shifts[(e, d, s)] = model.NewBoolVar(f'shift_e{e}_d{d}_s{s}')
        
        # 制約
        for e in range(num_employees):
            for d in range(num_days):
                model.Add(sum(shifts[(e, d, s)] for s in range(num_shifts)) == 1)
        
        for d in range(num_days):
            model.Add(sum(shifts[(e, d, 0)] for e in range(num_employees)) >= 1)
            model.Add(sum(shifts[(e, d, 1)] for e in range(num_employees)) >= 1)
            
        for e in range(num_employees):
            for d in range(num_days - 1):
                model.Add(shifts[(e, d, 1)] + shifts[(e, d + 1, 0)] <= 1)
                
        for e in range(num_employees):
            rest_count = sum(shifts[(e, d, 2)] for d in range(num_days))
            model.Add(rest_count >= 9)
            model.Add(rest_count <= 11)
            
        # 目的関数
        objective_terms = []
        for e, emp_name in enumerate(employees):
            for d in range(num_days):
                pref = preferences[emp_name][d]
                if pref == '早番': objective_terms.append(20 * shifts[(e, d, 0)])
                elif pref == '遅番': objective_terms.append(20 * shifts[(e, d, 1)])
                elif pref in ['希望休', '有給', '半休']: objective_terms.append(30 * shifts[(e, d, 2)])
                elif pref == 'どちらでも':
                    objective_terms.append(5 * shifts[(e, d, 0)])
                    objective_terms.append(5 * shifts[(e, d, 1)])
                    
        # 休み日数の平準化ボーナス
        for e in range(num_employees):
            rest_count = sum(shifts[(e, d, 2)] for d in range(num_days))
            model.Add(rest_count == self.target_rest_days).OnlyEnforceIf(model.NewBoolVar(f'obj_e{e}'))
            
        model.Maximize(sum(objective_terms))
        
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 60
        status = solver.Solve(model)
        
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            result = {}
            stats = {'rest_counts': {}, 'early_counts': {}, 'late_counts': {}}
            
            for e, emp_name in enumerate(employees):
                result[emp_name] = []
                for d in range(num_days):
                    for s in range(num_shifts):
                        if solver.Value(shifts[(e, d, s)]) == 1:
                            if s == 2:
                                orig = preferences[emp_name][d]
                                val = '有給' if orig == '有給' else ('半休' if orig == '半休' else '休み')
                                result[emp_name].append(val)
                            else:
                                result[emp_name].append(['早番', '遅番', '休み'][s])
                            break
                
                # 統計計算
                counts = result[emp_name]
                stats['early_counts'][emp_name] = counts.count('早番')
                stats['late_counts'][emp_name] = counts.count('遅番')
                stats['rest_counts'][emp_name] = counts.count('休み') + counts.count('有給') + (counts.count('半休') * 0.5)
                
            print("✅ 最適化完了！")
            return result, stats
        else:
            print("❌ 最適解が見つかりませんでした")
            return None, None

    def save_to_folder(self, output_file, employees, dates, preferences, result, stats):
        """
        確実な保存処理：一時ファイルを使用し、設定シート経由でプルダウンを作成
        """
        print(f"\n📝 結果の保存を開始します...")
        
        # 1. 一時ファイル名の作成
        temp_file = output_file.parent / f"temp_{int(time.time())}.xlsx"
        
        try:
            # 2. Pandasで基本データを書き込み
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                # シフト表
                data = []
                for emp in employees:
                    row = [emp] + result[emp]
                    data.append(row)
                cols = ['氏名'] + [d.strftime("%m/%d") for d in dates]
                pd.DataFrame(data, columns=cols).to_excel(writer, sheet_name='シフト表', index=False)
                
                # 休みカウント枠（データなし、枠のみ）
                pd.DataFrame([{'従業員': e} for e in employees]).to_excel(writer, sheet_name='休み日数カウント', index=False)
                
                # 設定シート（プルダウンのマスタ用）
                pd.DataFrame({'選択肢': ['早番', '遅番', '休み', '有給', '半休']}).to_excel(writer, sheet_name='設定', index=False)
                
                # その他シート
                self._create_calendar_view(writer, employees, dates, result)
                self._create_statistics_summary(writer, employees, stats, preferences, result, dates)
                self._create_comparison_sheet(writer, employees, dates, preferences, result)

            print("   - 基本データの作成完了")

            # 3. OpenPyXLで開いて機能を追加（プルダウン・数式・書式）
            wb = openpyxl.load_workbook(temp_file)
            
            # --- プルダウンメニューの設定（設定シート参照方式） ---
            ws_shift = wb['シフト表']
            ws_config = wb['設定']
            
            # 設定シートの範囲を定義（A2:A6）
            dv = DataValidation(type="list", formula1="'設定'!$A$2:$A$6", allow_blank=False, showDropDown=True)
            ws_shift.add_data_validation(dv)
            
            for r in range(2, len(employees) + 2):
                for c in range(2, len(dates) + 2):
                    dv.add(ws_shift.cell(row=r, column=c))
            
            print("   - プルダウンメニューの追加完了")
            
            # --- 数式と書式の追加 ---
            ws_count = wb['休み日数カウント']
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # ヘッダー
            headers = ['従業員', '早番', '遅番', '希望休', '有給', '半休', '合計休み日数']
            for i, h in enumerate(headers, 1):
                cell = ws_count.cell(row=1, column=i, value=h)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # 数式の埋め込み
            last_col = get_column_letter(len(dates) + 1)
            for i, emp in enumerate(employees):
                r = i + 2
                ws_count.cell(row=r, column=1, value=emp).border = thin_border
                
                range_str = f"シフト表!B{r}:{last_col}{r}"
                # 数式をセット
                ws_count.cell(row=r, column=2, value=f'=COUNTIF({range_str},"早番")').border = thin_border
                ws_count.cell(row=r, column=3, value=f'=COUNTIF({range_str},"遅番")').border = thin_border
                ws_count.cell(row=r, column=4, value=f'=COUNTIF({range_str},"休み")').border = thin_border
                ws_count.cell(row=r, column=5, value=f'=COUNTIF({range_str},"有給")').border = thin_border
                ws_count.cell(row=r, column=6, value=f'=COUNTIF({range_str},"半休")').border = thin_border
                ws_count.cell(row=r, column=7, value=f'=D{r}+E{r}+F{r}*0.5').border = thin_border

            print("   - 自動計算数式の追加完了")
            
            # 書式調整（シフト表）
            for row in ws_shift.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
            
            # 設定シートを隠す
            ws_config.sheet_state = 'hidden'

            # 4. 最終ファイルとして保存
            wb.save(output_file)
            wb.close()
            print(f"✅ 保存完了: {output_file}")
            
            # 一時ファイルの削除
            try:
                os.remove(temp_file)
            except:
                pass
                
            return output_file

        except Exception as e:
            print(f"\n❌ 保存中にエラーが発生しました: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _create_calendar_view(self, writer, employees, dates, result):
        calendar_data = []
        for d, date in enumerate(dates):
            day_info = {'日付': date.strftime("%Y/%m/%d"), '曜日': ['月','火','水','木','金','土','日'][date.weekday()]}
            early = [e for e in employees if result[e][d] == '早番']
            late = [e for e in employees if result[e][d] == '遅番']
            rest = [e for e in employees if result[e][d] == '休み']
            day_info['早番'] = ', '.join(early) if early else '-'
            day_info['遅番'] = ', '.join(late) if late else '-'
            day_info['休み'] = ', '.join(rest) if rest else '-'
            calendar_data.append(day_info)
        pd.DataFrame(calendar_data).to_excel(writer, sheet_name='カレンダー', index=False)

    def _create_statistics_summary(self, writer, employees, stats, preferences, result, dates):
        stat_data = []
        for emp in employees:
            stat_data.append({
                '従業員': emp,
                '早番日数': stats['early_counts'][emp],
                '遅番日数': stats['late_counts'][emp],
                '休み日数(合計)': stats['rest_counts'][emp]
            })
        pd.DataFrame(stat_data).to_excel(writer, sheet_name='統計', index=False)

    def _create_comparison_sheet(self, writer, employees, dates, preferences, result):
        comp_data = []
        for emp in employees:
            for d, date in enumerate(dates):
                pref = preferences[emp][d]
                act = result[emp][d]
                match = "○" if pref == 'どちらでも' or pref == act or (pref in ['希望休','有給','半休'] and act in ['休み','有給','半休']) else "×"
                comp_data.append({'従業員': emp, '日付': date.strftime("%m/%d"), '希望': pref, '実際': act, '一致': match})
        pd.DataFrame(comp_data).to_excel(writer, sheet_name='希望比較', index=False)

    def run(self):
        try:
            print("\n" + "=" * 60)
            print("🚀 シフト自動最適化プログラム Ver3.2")
            print("=" * 60)
            year_month = self.get_year_month_input()
            folder_path, input_file, output_file = self.setup_file_paths(year_month)
            employees, preferences, dates = self.load_excel_from_folder(input_file, year_month)
            result, stats = self.optimize_shifts(employees, preferences, len(dates))
            if result:
                self.save_to_folder(output_file, employees, dates, preferences, result, stats)
                # PADから実行された場合はGUIメッセージを出さない（処理が止まるのを防ぐため）
                if GUI_AVAILABLE and len(sys.argv) <= 1:
                    try:
                        root = tk.Tk(); root.withdraw()
                        messagebox.showinfo("完了", f"完了しました！\n{output_file}")
                        root.destroy()
                    except: pass
            else:
                print("❌ 最適解が見つかりませんでした")
                sys.exit(1)
        except Exception as e:
            print(f"\n❌ エラーが発生しました: {str(e)}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

if __name__ == "__main__":
    ShiftOptimizerV3_2(base_path="output").run()